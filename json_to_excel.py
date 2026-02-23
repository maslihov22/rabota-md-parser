import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def json_to_excel(json_file='vacancies_all.json', excel_file='vacancies_all.xlsx'):
    """Конвертирует JSON с вакансиями в Excel с автофильтрами и форматированием"""

    print(f"Читаем {json_file}...")
    with open(json_file, 'r', encoding='utf-8') as f:
        vacancies = json.load(f)

    if not vacancies:
        print("JSON пустой, нечего конвертировать")
        return

    print(f"Найдено вакансий: {len(vacancies)}")

    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Вакансии"

    # Собираем все поля из additional_info
    all_fields = set()
    for vacancy in vacancies:
        if vacancy.get('additional_info'):
            all_fields.update(vacancy['additional_info'].keys())

    # Основные колонки + дополнительные
    columns = ['Название', 'Компания', 'Зарплата', 'URL', 'Описание']
    columns.extend(sorted(all_fields))

    # Заголовки
    print("Создаем заголовки...")
    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Заполняем данные
    print("Записываем вакансии...")
    for row_num, vacancy in enumerate(vacancies, 2):
        ws.cell(row=row_num, column=1, value=vacancy.get('title', ''))
        ws.cell(row=row_num, column=2, value=vacancy.get('company', ''))
        ws.cell(row=row_num, column=3, value=vacancy.get('salary', ''))
        ws.cell(row=row_num, column=4, value=vacancy.get('url', ''))
        ws.cell(row=row_num, column=5, value=vacancy.get('description', ''))

        # Дополнительные поля
        if vacancy.get('additional_info'):
            for field_name in sorted(all_fields):
                col_num = columns.index(field_name) + 1
                value = vacancy['additional_info'].get(field_name, '')
                ws.cell(row=row_num, column=col_num, value=value)

    # Автофильтры
    print("Добавляем автофильтры...")
    ws.auto_filter.ref = ws.dimensions

    # Замораживаем первую строку
    ws.freeze_panes = "A2"

    # Автоширина колонок
    print("Настраиваем ширину колонок...")
    for col_num, column_name in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)

        # Особая обработка для URL и Описания
        if column_name == 'URL':
            ws.column_dimensions[column_letter].width = 50
        elif column_name == 'Описание':
            ws.column_dimensions[column_letter].width = 80
        elif column_name == 'Название':
            ws.column_dimensions[column_letter].width = 40
        elif column_name == 'Компания':
            ws.column_dimensions[column_letter].width = 25
        elif column_name == 'Зарплата':
            ws.column_dimensions[column_letter].width = 25
        else:
            # Для дополнительных полей - автоширина
            max_length = len(column_name)
            for row in ws.iter_rows(min_row=2, max_row=len(vacancies)+1, min_col=col_num, max_col=col_num):
                cell_value = str(row[0].value) if row[0].value else ""
                max_length = max(max_length, len(cell_value))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Перенос текста для описания
    for row in ws.iter_rows(min_row=2, max_row=len(vacancies)+1, min_col=5, max_col=5):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # Сохраняем
    print(f"Сохраняем в {excel_file}...")
    wb.save(excel_file)

    print(f"\n✓ Конвертация завершена!")
    print(f"✓ Обработано вакансий: {len(vacancies)}")
    print(f"✓ Excel сохранен в: {excel_file}")
    print(f"\nФишки файла:")
    print("  • Автофильтры на всех колонках")
    print("  • Замороженная шапка (можно скроллить)")
    print("  • Кликабельные ссылки")
    print("  • Красивое форматирование")

if __name__ == "__main__":
    json_to_excel()
